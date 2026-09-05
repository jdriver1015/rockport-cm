"""
Extracts the Aston construction tracker workbook to JSON for load-aston-tracker.ts.

Two stages rather than one because ExcelJS does not surface a cached result for
every formula cell in this file — CM/Supervision and Contingency (percentage
formulas) came back empty, which silently dropped the two largest add-ons from
each renovation tier. openpyxl with data_only=True reads all of them, so the
extraction happens here and the loading stays in TypeScript against the schema.

  python scripts/extract-aston-tracker.py <workbook.xlsx> <out.json>
"""
import json
import sys

import openpyxl


def s(v):
    if v is None:
        return None
    t = str(v).strip()
    return t or None


def n(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def main(src: str, dest: str) -> None:
    wb = openpyxl.load_workbook(src, data_only=True)
    out = {}

    # ---- non-interior capex: leaf rows only -------------------------------
    ws = wb["Rockport CapEX Budget"]
    subtotals = {
        "total", "grand total", "exterior and common areas",
        "unit renovations", "scope description", "scope",
    }
    capex = []
    for r in range(3, 64):
        desc, amt = s(ws.cell(r, 2).value), n(ws.cell(r, 3).value)
        if not desc or amt is None or desc.lower() in subtotals:
            continue
        capex.append({
            "item": desc,
            "amount": amt,
            "category": None,
            "awardedVendor": s(ws.cell(r, 8).value),
            "task": s(ws.cell(r, 9).value),
            "priority": s(ws.cell(r, 10).value),
            "status": s(ws.cell(r, 11).value),
            "update": s(ws.cell(r, 12).value),
            "bids": [x for x in (n(ws.cell(r, c).value) for c in (5, 6, 7)) if x],
        })
    out["capex"] = capex

    # ---- interior tiers ----------------------------------------------------
    ws = wb["Rockport Interior CapEx Budget"]
    tiers = []
    for i, col in enumerate(range(2, 8)):
        name = s(ws.cell(2, col).value)
        lines = []
        for r in range(4, 20):
            scope, price = s(ws.cell(r, 1).value), n(ws.cell(r, col).value)
            if scope and price:
                lines.append({"scope": scope, "price": price})
        tiers.append({
            "name": name,
            "plannedUnitsRaw": n(ws.cell(3, col).value),
            "perUnitTotal": n(ws.cell(20, col).value),
            "lines": lines,
        })
    out["tiers"] = tiers

    # ---- the one turn in flight, with its real vendor pricing --------------
    # Rows 31-46 of the same sheet are actual costs booked to date; row 29
    # says which tier the completed unit belongs to.
    actual = []
    for r in range(31, 47):
        scope = s(ws.cell(r, 1).value)
        for col in range(2, 8):
            price = n(ws.cell(r, col).value)
            if scope and price:
                actual.append({"scope": scope, "price": price, "tier": s(ws.cell(28, col).value)})
    out["turnActuals"] = actual

    # ---- unit tracker ------------------------------------------------------
    ws = wb["Unit Tracker"]
    turns = []
    for r in range(7, ws.max_row + 1):
        plan, unit = s(ws.cell(r, 1).value), s(ws.cell(r, 2).value)
        if not plan or not unit:
            continue
        start, comp = ws.cell(r, 3).value, ws.cell(r, 4).value
        turns.append({
            "floorplan": plan,
            "unitNumber": unit,
            "start": start.strftime("%Y-%m-%d") if hasattr(start, "strftime") else None,
            "complete": comp.strftime("%Y-%m-%d") if hasattr(comp, "strftime") else None,
            "status": s(ws.cell(r, 6).value),
            "previousRent": n(ws.cell(r, 12).value),
            "tradeOutRent": n(ws.cell(r, 13).value),
        })
    out["turns"] = turns

    # ---- rent roll ---------------------------------------------------------
    ws = wb["Rent Roll"]
    units = []
    for r in range(3, ws.max_row + 1):
        u = s(ws.cell(r, 1).value)
        if not u or u.lower() == "unit no.":
            continue
        units.append({
            "unitNumber": u,
            "floorPlanCode": s(ws.cell(r, 2).value),
            "sqft": n(ws.cell(r, 3).value),
            "beds": n(ws.cell(r, 4).value),
            "baths": n(ws.cell(r, 5).value),
            "renovated": (s(ws.cell(r, 7).value) or "").lower() == "renovated",
            "occupancy": s(ws.cell(r, 8).value),
            "marketRent": n(ws.cell(r, 9).value),
            "inPlaceRent": n(ws.cell(r, 10).value),
        })
    out["rentRoll"] = {"asOfDate": "2026-06-30", "units": units}

    with open(dest, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=1)

    print(f"capex leaf lines : {len(capex)}  (${sum(c['amount'] for c in capex):,.2f})")
    print(f"tiers            : {len(tiers)}")
    for t in tiers:
        got = sum(l["price"] for l in t["lines"])
        flag = "" if t["perUnitTotal"] and abs(got - t["perUnitTotal"]) < 0.01 else "  <-- MISMATCH"
        print(f"   {t['name']:20} {len(t['lines']):2} lines  ${got:>10,.2f}  sheet ${t['perUnitTotal'] or 0:>10,.2f}{flag}")
    print(f"turn actuals     : {len(actual)} line(s)")
    print(f"unit turns       : {len(turns)}")
    print(f"rent roll units  : {len(units)}  ({sum(1 for u in units if u['renovated'])} renovated)")
    print(f"awarded vendors  : {len({c['awardedVendor'] for c in capex if c['awardedVendor']})}")
    print(f"\nwrote {dest}")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
